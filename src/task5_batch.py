"""
Task 5 — Batch Processing & Report Generation [30 points]
Processes multiple quiz images and generates a structured Excel/CSV report.
"""

import os
import sys
import glob
import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Optional

# Add parent directory to path when running standalone
sys.path.insert(0, os.path.dirname(__file__))

from task1_qr_decoder import decode_answer_key, AnswerKey
from task2_ocr import extract_student_info
from task3_bubble_reader import read_bubble_sheet
from task4_grader import grade_quiz, GradeReport

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ── Single-image pipeline ─────────────────────────────────────────────────────

def process_single_image(image_path: str, quiz_label: str = "Quiz 1",
                          fallback_key: Optional[AnswerKey] = None) -> Optional[dict]:
    """
    Run the full pipeline (Tasks 1–4) on one image.
    Returns a flat dict of all columns for the report, or None on critical failure.
    """
    import cv2
    print(f"\n{'─'*55}")
    print(f"Processing: {image_path}")
    print(f"{'─'*55}")

    image = cv2.imread(image_path)
    if image is None:
        print(f"  ERROR: Cannot read image file.")
        return None

    # Task 1
    answer_key = decode_answer_key(image)
    if answer_key is None:
        if fallback_key is not None:
            print("  Using provided fallback answer key.")
            answer_key = fallback_key
        else:
            print("  ERROR: No QR code found and no fallback key — skipping.")
            return None

    # Task 2
    student_info = extract_student_info(image)

    # Task 3
    student_answers = read_bubble_sheet(image)

    # Task 4
    report = grade_quiz(
        student_answers, answer_key,
        student_name=student_info.name,
        reg_no=student_info.reg_no,
    )

    # Build flat row
    row = {
        "Quiz": quiz_label,
        "Set": answer_key.quiz_set,
        "Class": answer_key.class_name or "BSE-4A",
        "Subject": answer_key.subject or "Artificial Intelligence",
        "Name": report.student_name,
        "Reg No": report.reg_no,
    }

    # Part-I answers
    for q in [f"Q{i:02d}" for i in range(1, 9)]:
        row[f"Part1_{q}"] = student_answers.part1.get(q) or ""

    # Part-II answers
    for q in [f"Q{i:02d}" for i in range(1, 9)]:
        row[f"Part2_{q}"] = student_answers.part2.get(q) or ""

    row["Correct"]      = report.correct
    row["Incorrect"]    = report.incorrect
    row["Unattempted"]  = report.unattempted
    row["Total Marks"]  = report.total_marks
    row["Max Marks"]    = report.max_marks
    row["Percentage"]   = round(report.percentage, 2)
    row["Grade"]        = report.grade

    return row


# ── Batch runner ─────────────────────────────────────────────────────────────

def run_batch(image_dir: str, quiz_label: str = "Quiz 1",
              fallback_key: Optional[AnswerKey] = None,
              output_dir: str = "output") -> str:
    """
    Process all images in image_dir.
    Returns path to the generated report file.
    """
    os.makedirs(output_dir, exist_ok=True)

    # Collect images
    extensions = ["*.jpg", "*.jpeg", "*.png", "*.bmp", "*.tiff"]
    image_paths = []
    for ext in extensions:
        image_paths.extend(glob.glob(os.path.join(image_dir, ext)))
        image_paths.extend(glob.glob(os.path.join(image_dir, ext.upper())))
    image_paths = sorted(set(image_paths))

    if not image_paths:
        raise ValueError(f"No image files found in: {image_dir}")

    print(f"\n{'='*55}")
    print(f"  BATCH PROCESSING — {len(image_paths)} images")
    print(f"{'='*55}")

    rows = []
    for i, path in enumerate(image_paths, 1):
        row = process_single_image(path, quiz_label=quiz_label, fallback_key=fallback_key)
        if row:
            rows.append(row)

    if not rows:
        raise ValueError("No results produced — check your images.")

    # Auto-named output file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_label = quiz_label.replace(" ", "_")
    base_name = f"{safe_label}_{timestamp}"

    if OPENPYXL_AVAILABLE:
        out_path = os.path.join(output_dir, base_name + ".xlsx")
        write_excel(rows, out_path, quiz_label)
    else:
        out_path = os.path.join(output_dir, base_name + ".csv")
        write_csv(rows, out_path)

    print(f"\n{'='*55}")
    print(f"  Report saved: {out_path}")
    print(f"{'='*55}")
    return out_path


# ── CSV writer ────────────────────────────────────────────────────────────────

def write_csv(rows: list[dict], path: str):
    if not rows:
        return
    fieldnames = list(rows[0].keys())
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

        # Summary row
        avg_pct    = sum(r["Percentage"] for r in rows) / len(rows)
        max_marks  = max(r["Total Marks"] for r in rows)
        min_marks  = min(r["Total Marks"] for r in rows)
        summary = {k: "" for k in fieldnames}
        summary["Name"]        = "SUMMARY"
        summary["Percentage"]  = round(avg_pct, 2)
        summary["Total Marks"] = f"Avg:{avg_pct:.1f} Max:{max_marks} Min:{min_marks}"
        writer.writerow(summary)

    print(f"  CSV written: {path}")


# ── Excel writer ──────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F3864") if OPENPYXL_AVAILABLE else None
CORRECT_FILL = PatternFill("solid", fgColor="C6EFCE") if OPENPYXL_AVAILABLE else None
WRONG_FILL   = PatternFill("solid", fgColor="FFC7CE") if OPENPYXL_AVAILABLE else None
ALT_FILL     = PatternFill("solid", fgColor="EEF2FF") if OPENPYXL_AVAILABLE else None
SUMMARY_FILL = PatternFill("solid", fgColor="FFE699") if OPENPYXL_AVAILABLE else None


def _border():
    thin = Side(style="thin", color="BBBBBB")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def write_excel(rows: list[dict], path: str, quiz_label: str = "Quiz"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    if not rows:
        wb.save(path)
        return

    headers = list(rows[0].keys())
    ws.append(headers)

    # Style header row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        is_alt = row_idx % 2 == 0
        for col_idx, key in enumerate(headers, 1):
            val = row[key]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _border()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(size=9)
            # Colour grade column
            if key == "Grade":
                if val in ("A+", "A", "A-"):
                    cell.fill = CORRECT_FILL
                elif val == "F":
                    cell.fill = WRONG_FILL
                elif is_alt:
                    cell.fill = ALT_FILL
            elif is_alt:
                cell.fill = ALT_FILL

    # Summary row
    summary_row_idx = len(rows) + 2
    avg_pct   = sum(r["Percentage"] for r in rows) / len(rows)
    max_marks = max(r["Total Marks"] for r in rows)
    min_marks = min(r["Total Marks"] for r in rows)

    summary_values = {k: "" for k in headers}
    summary_values["Name"]        = "CLASS SUMMARY"
    summary_values["Total Marks"] = round(avg_pct / 100 * (rows[0]["Max Marks"] if rows else 16), 2)
    summary_values["Percentage"]  = round(avg_pct, 2)
    summary_values["Grade"]       = f"Avg {avg_pct:.1f}%  Max {max_marks}  Min {min_marks}"

    for col_idx, key in enumerate(headers, 1):
        cell = ws.cell(row=summary_row_idx, column=col_idx, value=summary_values[key])
        cell.fill = SUMMARY_FILL
        cell.font = Font(bold=True, size=9)
        cell.border = _border()
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-fit column widths
    for col_idx, header in enumerate(headers, 1):
        max_len = max(
            len(str(header)),
            *[len(str(r[header])) for r in rows],
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 25)

    # Freeze header
    ws.freeze_panes = "A2"

    # Summary stats sheet
    ws2 = wb.create_sheet("Summary Stats")
    ws2.append(["Metric", "Value"])
    ws2.append(["Total Students", len(rows)])
    ws2.append(["Class Average (%)", round(avg_pct, 2)])
    ws2.append(["Highest Score", max_marks])
    ws2.append(["Lowest Score", min_marks])
    grade_counts = {}
    for r in rows:
        grade_counts[r["Grade"]] = grade_counts.get(r["Grade"], 0) + 1
    for g, count in sorted(grade_counts.items()):
        ws2.append([f"Grade {g}", count])

    wb.save(path)
    print(f"  Excel written: {path}")


# ── CLI entry point ───────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Batch Quiz Scanner")
    parser.add_argument("image_dir", help="Folder containing quiz images")
    parser.add_argument("--quiz", default="Quiz 1", help="Quiz label")
    parser.add_argument("--output", default="output", help="Output directory")
    parser.add_argument("--fallback-key", default=None,
                        help="JSON string of fallback answer key if QR fails")
    args = parser.parse_args()

    fallback = None
    if args.fallback_key:
        data = json.loads(args.fallback_key)
        fallback = AnswerKey(**data)

    out = run_batch(args.image_dir, quiz_label=args.quiz,
                    fallback_key=fallback, output_dir=args.output)
    print(f"\nDone! Report: {out}")
